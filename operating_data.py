from openpyxl import load_workbook, Workbook



x = []
def get_data():
     print()
     print('Запущена программа обработки данных БУР, в которой программа уберет все промежуточные значения'
           'и позволит использовать выходящие данные в программе товарища Хильченко для расчета коэффициента'
           'расхода ресурса. Пожалуйста, вводите данные с предельной точностью. Файл с данными должен находиться'
           'в одной директории с исполняемой программой. Данные для обработки должны быть записаны в одну колонку.'
           'dndia v.1.0')
     print()
     print()
     name_wb = input('Введите имя файла Exel, в котором содержатся входящие данные (с раширением): ')
     name_ws = input('Введите имя рабочего листа с входящими данными: ')
     col_range_start = input('Введите первую ячейку в колонке с данными: ')
     col_range_finish = input('Введите крайнюю ячейку в колонке с данными: ')
     wb = load_workbook(filename = name_wb, data_only = True)
     ws = wb[name_ws]
     cell_range = ws[col_range_start:col_range_finish]
     for row in cell_range:
          for i in row:
               x.append(i.value)
     print(x)
     return x
get_data()

def operate_data(x):
     tech =[]
     a = 0
     tech.append(x[a])
     print()
     print('on start append ' + str(x[a]))
     while a <= len(x):
               try:
                    print(a)
                    if x[a] < x[a+1] or x[a] == x[a+1]:
                         print()
                         print('more in ' + str(a))
                         b = int(a)
                         try:
                              while x[b] < x[b+1] or x[b] == x[b+1]:
                                   print(str(x[b+1]) + ' more ' + str(x[b]))
                                   b += 1
                         except IndexError:
                              pass
                         tech.append(x[b])
                         print('append ' + str(x[b]))
                         a = b
                         print('out ' + str(a))
                    elif x[a] > x[a+1] or x[a] == x[a+1]:
                         print()
                         print('low in ' + str(a))
                         c = int(a)
                         try:
                              while x[c] > x[c+1] or x[c] == x[c+1]:
                                   print(str(x[c + 1]) + ' low ' + str(x[c]))
                                   c += 1
                         except IndexError:
                              pass
                         tech.append(x[c])
                         a = c
                         print('out ' + str(a))
               except IndexError:
                    a = len(x)+1
     print(tech)
     return tech

tech = operate_data(x)

def push_data(tech):
     print()
     print()
     name_wb = input('Введите имя файла Exel, в котором будут содержаться обработанные данные (без расширения): ')
     col_range = input('Введите имя колонки (большая латинская буква), в которую поместить данные: ')
     str_range = input('Введите начальную строку колонки: ')
     wb = Workbook()
     ws = wb.create_sheet("обработанные данные")
     ws.title = "обработанные данные"
     colum_word = col_range
     str_number = str_range
     str_number = int(str_number)
     for i in tech:
          ws[colum_word+str(str_number)] = i
          str_number +=1
     wb.save(name_wb + '.xlsx')
     print()
     print('Данные готовы.')

push_data(tech)





