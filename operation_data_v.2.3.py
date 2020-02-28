import PySimpleGUI as sg
from openpyxl import load_workbook, Workbook
import os

sg.popup('Информация', 'Запущена программа обработки данных БУР, работающая в связке с программой к.т.н. Хильченко. Убирает все промежуточные значения в данных. \n'
                       '\n'
                       'dndia v.2.3')


layout = [
    [sg.Text('Файл с данными:', size=(13, 1)), sg.InputText('Только файлы с расширением ".xlsx"', key='wb'), sg.FileBrowse('Выбрать файл'),],
    [sg.Text('Рабочий лист:', size=(13, 1)), sg.InputText('Лист1', key='ws'),],
    [sg.Text('Буква колонки:', size=(13, 1)), sg.Drop(values=('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'), auto_size_text=True, size=(15, 1), key='word')],
    [
        sg.Text('Начальная строка:', size=(15, 1)), sg.Spin(values=[i for i in range(1, 1000)], initial_value=1, size=(6, 3), key='f_num'),
        sg.Text('Конечная строка:', size=(15, 1)), sg.Spin(values=[i for i in range(1, 1000)], initial_value=330, size=(6, 1), key='s_num'),
        ],
    [sg.Text('Путь сохранения:', size=(13, 1)), sg.InputText('', key='path'), sg.FolderBrowse('Выбрать папку')],
    [
        sg.Text('Буква колонки сохранения:', size=(21, 1)), sg.Drop(values=('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'), auto_size_text=True, size=(15, 1), key='word_save'),
        sg.Text('Начальная строка сохранения:', size=(22, 1)), sg.Spin(values=[i for i in range(1, 1000)], initial_value=3, size=(6, 1), key='f_num_save'),
        ],
    [sg.Text('Cохраняемое расширение файла:', size=(29, 1)), sg.Checkbox('.xlsx', size=(5,1), default=True, key='xlsx'),  sg.Checkbox('.txt', default=True, key='txt')],
    [sg.Text('Префикс к имени сохраняемого файла:', size=(29, 1)), sg.InputText('_result_', key='pref')],
    [sg.Output(size=(88, 20))],
    [sg.Submit('Начать обработку'), sg.Cancel('Выход')]
]
window = sg.Window('Operation Data', layout)
print('Если программа не отвечает после нажатия кнопки "Началь обработку" - не нужно пугаться, просто идет обработка данных. \n'
      'Заварите чайку, выпейте...')
while True:                             # The Event Loop
    event, values = window.read()
    # print(event, values) #debug
    if event in (None, 'Выход', 'Cancel'):
        break
    if event == 'Начать обработку':
        sg.popup('Начинается обработка. Если программа не отвечает - не нужно пугаться, просто идет обработка данных. \n'
                 'Заварите чайку, выпейте...\n'
                 '\n'
                 'Но если вдруг программа просто сама закрылась и результат не выдан - вы где-то совершили ошибку. \n'
                 'Запустите программу еще раз, введите данные с большей точностью.')

        print('Обработка данных. Пожалуйста, подождите...')
        x = []


        def get_data():
            print()
            print()
            print()
            name_wb = values['wb']
            name_ws = values['ws']
            col_range_start = values['word']+str(values['f_num'])
            col_range_finish = values['word']+str(values['s_num'])
            wb = load_workbook(filename=name_wb, data_only=True)
            ws = wb[name_ws]
            cell_range = ws[col_range_start:col_range_finish]
            for row in cell_range:
                for i in row:
                    x.append(i.value)
            print(x)
            return x


        get_data()


        def operate_data(x):
            tech = []
            a = 0
            tech.append(x[a])
            print()
            print('on start append ' + str(x[a]))
            while a <= len(x):
                try:
                    print(a)
                    if x[a] < x[a + 1] or x[a] == x[a + 1]:
                        print()
                        print('more in ' + str(a))
                        b = int(a)
                        try:
                            while x[b] < x[b + 1] or x[b] == x[b + 1]:
                                print(str(x[b + 1]) + ' more ' + str(x[b]))
                                b += 1
                        except IndexError:
                            pass
                        tech.append(x[b])
                        print('append ' + str(x[b]))
                        a = b
                        print('out ' + str(a))
                    elif x[a] > x[a + 1] or x[a] == x[a + 1]:
                        print()
                        print('low in ' + str(a))
                        c = int(a)
                        try:
                            while x[c] > x[c + 1] or x[c] == x[c + 1]:
                                print(str(x[c + 1]) + ' low ' + str(x[c]))
                                c += 1
                        except IndexError:
                            pass
                        tech.append(x[c])
                        a = c
                        print('out ' + str(a))
                except IndexError:
                    a = len(x) + 1
            print(tech)
            return tech


        tech = operate_data(x)


        def push_data(tech):
            print()
            print()
            name_wb = os.path.basename(values['wb'])[:-5] + values['pref']
            col_range = values['word_save']
            str_range = values['f_num_save']

            if values['xlsx'] == True:
                wb = Workbook()
                ws = wb.create_sheet("обработанные данные", 0)
                ws.title = "обработанные данные"
                colum_word = col_range
                str_number = str_range
                str_number = int(str_number)
                for i in tech:
                    ws[colum_word + str(str_number)] = i
                    str_number += 1
                wb.save(values['path'] + '/' + name_wb + '.xlsx')
                print()
                print('Обработанные данные формата .xlsx сохранены по адресу:')
                print(values['path'] + '/' + name_wb + '.xlsx')

            if values['txt'] == True:
                f = open(values['path'] + '/' + name_wb + '.txt', 'w')
                for i in tech:
                    f.write(str(i) + '\n')
                f.close()
                print()
                print('Обработанные данные формата .txt сохранены по адресу:')
                print(values['path'] + '/' + name_wb + '.xlsx')

            print()
            print('Работа с текущей базой данных завершена.')



        push_data(tech)
